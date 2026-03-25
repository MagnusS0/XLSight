"""Generate realistic .xlsx test fixtures for XLSight.

Files produced (all written to the same directory as this script):
  small.xlsx       — 10 rows × 5 cols, all cell types, named range, merge
  medium.xlsx      — 1 000 rows × 10 cols, numbers + text + dates + booleans
  large.xlsx       — 100 000 rows × 5 cols of numbers  (streaming benchmark)
  string_heavy.xlsx— 2 000 rows × 5 cols, shared-string-heavy text data
  named_ranges.xlsx— named ranges, merged cells, multiple sheets
"""

import os
import random
import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent
random.seed(42)


# ---------------------------------------------------------------------------
# small.xlsx — all cell types, named range "SalesData", one merge
# ---------------------------------------------------------------------------
def make_small():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["ID", "Name", "Score", "Active", "Date"]
    ws.append(headers)

    names = ["Alice", "Bob", "Carol", "Dave", "Eve",
             "Frank", "Grace", "Hank", "Iris", "Jack"]
    base = datetime.date(2024, 1, 1)

    for i, name in enumerate(names, start=1):
        ws.append([
            i,
            name,
            round(50 + random.random() * 50, 2),
            bool(i % 2),
            base + datetime.timedelta(days=i * 30),
        ])

    # Named range covering the data (excluding header)
    wb.defined_names["SalesData"] = openpyxl.workbook.defined_name.DefinedName(
        name="SalesData", attr_text="Sheet1!$A$2:$E$11"
    )

    # Merge A12:E12 as a footer
    ws.append(["", "", "", "", "End of data"])
    ws.merge_cells("A12:D12")

    # Second empty sheet
    wb.create_sheet("EmptySheet")

    wb.save(OUT / "small.xlsx")
    print("  small.xlsx — 10 data rows, all types, named range, merge")


# ---------------------------------------------------------------------------
# medium.xlsx — 1 000 rows × 10 cols, mixed types
# ---------------------------------------------------------------------------
def make_medium():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["RowID", "Category", "SubCat", "Value1", "Value2",
               "Value3", "Flag", "Date", "Label", "Score"]
    ws.append(headers)

    categories = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon"]
    subcats = ["X1", "X2", "X3", "Y1", "Y2"]
    base = datetime.date(2023, 1, 1)

    for i in range(1, 1001):
        ws.append([
            i,
            random.choice(categories),
            random.choice(subcats),
            round(random.uniform(0, 10000), 2),
            round(random.uniform(-500, 500), 4),
            random.randint(1, 100),
            bool(i % 3 != 0),
            base + datetime.timedelta(days=i),
            f"Label_{i:04d}",
            round(random.gauss(50, 15), 1),
        ])

    wb.save(OUT / "medium.xlsx")
    print("  medium.xlsx — 1 000 rows × 10 cols")


# ---------------------------------------------------------------------------
# large.xlsx — 100 000 rows × 5 cols, numbers only  (streaming benchmark)
# ---------------------------------------------------------------------------
def make_large():
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Numbers")

    ws.append(["A", "B", "C", "D", "E"])  # header

    for i in range(1, 100_001):
        ws.append([
            i,
            round(random.uniform(0, 1_000_000), 2),
            round(random.uniform(-100, 100), 4),
            random.randint(0, 9999),
            round(random.gauss(0, 1), 6),
        ])

    wb.save(OUT / "large.xlsx")
    print("  large.xlsx — 100 000 rows × 5 cols numbers")


# ---------------------------------------------------------------------------
# string_heavy.xlsx — 2 000 rows, many repeated strings → large SST
# ---------------------------------------------------------------------------
def make_string_heavy():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Strings"

    headers = ["Region", "Product", "Channel", "Status", "Notes"]
    ws.append(headers)

    regions  = [f"Region_{c}" for c in "ABCDEFGHIJ"]
    products = [f"Product_{i:03d}" for i in range(1, 51)]
    channels = ["Online", "Retail", "Wholesale", "Direct", "Partner"]
    statuses = ["Active", "Pending", "Closed", "Cancelled", "Review"]

    for _ in range(2000):
        ws.append([
            random.choice(regions),
            random.choice(products),
            random.choice(channels),
            random.choice(statuses),
            f"Note {random.randint(1, 500)}",
        ])

    wb.save(OUT / "string_heavy.xlsx")
    print("  string_heavy.xlsx — 2 000 rows, high SST reuse")


# ---------------------------------------------------------------------------
# named_ranges.xlsx — two sheets, named ranges, merged cells
# ---------------------------------------------------------------------------
def make_named_ranges():
    wb = openpyxl.Workbook()

    # Sheet 1: quarterly summary
    ws1 = wb.active
    ws1.title = "Summary"

    quarters = ["Q1", "Q2", "Q3", "Q4"]
    ws1.append(["Quarter", "Revenue", "Cost", "Profit"])
    for i, q in enumerate(quarters, start=1):
        rev = round(random.uniform(100_000, 500_000), 0)
        cost = round(rev * random.uniform(0.4, 0.7), 0)
        ws1.append([q, rev, cost, rev - cost])

    # Merge title row spanning all columns
    ws1.insert_rows(1)
    ws1["A1"] = "Annual Financial Summary"
    ws1.merge_cells("A1:D1")

    # Named ranges (workbook-scoped)
    wb.defined_names["Revenue"] = openpyxl.workbook.defined_name.DefinedName(
        name="Revenue", attr_text="Summary!$B$3:$B$6"
    )
    wb.defined_names["Profit"] = openpyxl.workbook.defined_name.DefinedName(
        name="Profit", attr_text="Summary!$D$3:$D$6"
    )

    # Sheet 2: detail
    ws2 = wb.create_sheet("Detail")
    ws2.append(["Month", "Amount", "Category"])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    cats = ["Revenue", "Cost", "Tax"]
    for m in months:
        ws2.append([m, round(random.uniform(10_000, 50_000), 0),
                    random.choice(cats)])

    # Sheet-scoped named range
    scope_id = wb.sheetnames.index("Detail")
    wb.defined_names["MonthlyData"] = openpyxl.workbook.defined_name.DefinedName(
        name="MonthlyData",
        attr_text="Detail!$A$2:$C$13",
        localSheetId=scope_id,
    )

    # Merged cells in Detail
    ws2.merge_cells("A14:C14")
    ws2["A14"] = "Total"

    wb.save(OUT / "named_ranges.xlsx")
    print("  named_ranges.xlsx — 2 sheets, 3 named ranges, merged cells")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Generating XLSight test fixtures...")
    make_small()
    make_medium()
    make_large()
    make_string_heavy()
    make_named_ranges()
    print("Done.")
